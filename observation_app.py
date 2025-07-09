import os
import datetime
import base64
import io

import dash
from dash import dcc, html, Input, Output, State, no_update, ALL
import dash_bootstrap_components as dbc
from dash.exceptions import PreventUpdate
from flask_login import current_user

from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import db_observations
from ai_module import get_ai_analysis
import design_components
import user_management
from utils import compress_image # FIX: Import from utils

def create_observation_form_layout():
    INDUSTRY_LIST = [
        "General", "Oil", "Gas", "Construction", "Manufacturing", "Hospitality",
        "Healthcare", "Logistics", "Transportation", "Mining", "Quarrying",
        "Utilities", "Energy", "Nuclear", "Agriculture", "Food Processing",
        "Education", "Public Institutions", "Retail", "Commercial", "Marine",
        "Offshore", "Waste Management", "Recycling", "Chemicals", "Pharmaceuticals",
        "Power Generation"
    ]
    rating_options = [{'label': str(i), 'value': i} for i in range(1, 6)]
    return html.Div(className="observation-form-container", style={'maxWidth': '800px', 'margin': 'auto'}, children=[
        html.Div(id='industry-update-status'),
        html.Div(id='flash-messages-container', className="flash-messages"),
        html.Div(className="form-content", children=[
            html.Div([
                html.H1("Safety Observation Assistant", className="form-title", style={'marginBottom': '0'}),
                dcc.Link(dbc.Button("View Report", color="info"), href="/report")
            ], style={'display': 'flex', 'justifyContent': 'space-between', 'alignItems': 'center', 'marginBottom': '35px'}),
            dbc.Checklist(options=[{"label": "AI", "value": 1}], value=[1], id="ai-toggle-switch", switch=True, className="mb-4"),
            html.Div(className="form-group", children=[
                html.Label("Select Your Industry:", htmlFor="industry-dropdown"),
                dcc.Dropdown(id='industry-dropdown', options=[{'label': i, 'value': i} for i in INDUSTRY_LIST], value=current_user.industry, clearable=False)
            ]),
            html.Div(className="form-group", children=[
                html.Label("Area / Equipment:", htmlFor="area-equipment-input"),
                dcc.Input(type="text", id="area-equipment-input", placeholder="e.g., Conveyor Belt #3, Main Warehouse Forklift...")
            ]),
            html.Div(className="form-group", children=[
                html.Label("Observation Details:", id="observation-label", htmlFor="observation-textarea"),
                dcc.Textarea(id="observation-textarea", placeholder="Describe what you observed...", rows=6)
            ]),
            html.Div(id='manual-input-fields', style={'display': 'none'}, children=[
                html.Div(className="form-group", children=[
                    html.Label("Impact:", htmlFor="manual-impact-input"),
                    dcc.Textarea(id="manual-impact-input", placeholder="Describe the potential impact...", rows=3)
                ]),
                html.Div(className="form-group", style={'display': 'flex', 'gap': '20px'}, children=[
                    html.Div(style={'flex': 1}, children=[
                        html.Label("Likelihood (1-5):", htmlFor="manual-likelihood-input"),
                        dcc.Dropdown(id='manual-likelihood-input', options=rating_options, placeholder="Select..."),
                    ]),
                    html.Div(style={'flex': 1}, children=[
                        html.Label("Severity (1-5):", htmlFor="manual-severity-input"),
                        dcc.Dropdown(id='manual-severity-input', options=rating_options, placeholder="Select..."),
                    ]),
                ]),
                html.Div(className="form-group", children=[
                    html.Label("Risk Rating (Likelihood x Severity):"),
                    dcc.Input(id="manual-risk-rating-input", type="number", readOnly=True, style={'backgroundColor': '#e9ecef'})
                ]),
                html.Div(className="form-group", children=[
                    html.Label("Corrective Action:", htmlFor="manual-corrective-action-input"),
                    dcc.Textarea(id="manual-corrective-action-input", placeholder="Describe the recommended action...", rows=3)
                ]),
                html.Div(className="form-group", children=[
                    html.Label("Deadline:", htmlFor="manual-deadline-input"),
                    dcc.Input(id="manual-deadline-input", type="text", placeholder="e.g., Immediately, 24 Hours, 1 Week...")
                ]),
            ]),
            html.Div(className="form-group", children=[
                html.Label("Attach Photo (Optional):"),
                dcc.Upload(id='photo-upload', className="photo-upload-area-styled", children=html.Div([
                    html.Img(src='/assets/upload-icon.png', className="upload-icon-img-styled"),
                    html.Span("Upload or Take Photo")
                ], className="photo-upload-button-styled"), multiple=False),
                html.Span(id="selected-file-name")
            ]),
            html.Div(className="submit-button-container", children=[
                # FIX: Add loading indicator to the submit button
                dcc.Loading(id="loading-add-observation", type="circle", children=[
                    html.Button("Add Observation to Database", id="add-button", n_clicks=0, className="submit-button-style")
                ])
            ])
        ])
    ])

def create_report_layout():
    return html.Div([
        dcc.Download(id='download-excel'),
        dcc.ConfirmDialog(id='confirm-delete-dialog', message='Are you sure you want to delete this observation?'),
        dcc.Store(id='store-id-to-delete'),
        dcc.Store(id='store-refresh-signal', data=0),
        html.Div(className="report-main-content", children=[
            html.H1("Your Safety Observation Report", className="form-title"),
            html.Div(id='delete-status-message'),
            html.Div(className="report-controls", children=[
                dcc.Input(id='search-input', type='text', placeholder='Search your reports by keyword or area...', debounce=True, className='search-bar'),
                dcc.Dropdown(
                    id='sort-dropdown',
                    options=[
                        {'label': 'Sort: Newest First', 'value': 'date_newest'},
                        {'label': 'Sort: Oldest First', 'value': 'date_oldest'},
                        {'label': 'Sort: Highest Risk', 'value': 'risk_high'},
                    ], value='date_newest', clearable=False, className='sort-dropdown-wrapper'
                ),
                dbc.Button("Download Report (XLSX)", id="download-report-button", color="success", className="nav-download-button"),
                dcc.Link(dbc.Button("New Observation", color="primary"), href="/observation", className="ms-2")
            ]),
            dcc.Loading(id="loading-report", type="default", children=html.Div(id='report-content-container'))
        ])
    ])

def generate_excel_for_download(observations_data):
    EXCEL_PHOTO_TARGET_WIDTH_PX = 150
    EXCEL_PHOTO_TARGET_HEIGHT_PX = 112
    EXCEL_ROW_HEIGHT_FOR_PHOTO_PT = 90.0
    EXCEL_PHOTO_COLUMN_WIDTH_UNITS = 22
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Safety Observation Report"
    green_fill = PatternFill(start_color="04A227", end_color="04A227", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFD406", end_color="FFD406", fill_type="solid")
    orange_fill = PatternFill(start_color="FF7A00", end_color="FF7A00", fill_type="solid")
    red_fill = PatternFill(start_color="F14219", end_color="F14219", fill_type="solid")
    sheet.merge_cells('A1:C4')
    logo_path = os.path.join('assets', 'riskwatch-logo.png')
    if os.path.exists(logo_path):
        logo = OpenpyxlImage(logo_path)
        logo.height = 80; logo.width = 180
        sheet.add_image(logo, 'A1')
    title_font = Font(name='Calibri', size=16, bold=True, color="000080")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    sheet.merge_cells('D1:N4')
    title_cell = sheet['D1']
    title_cell.value = "SAFETY OBSERVATION REPORT"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    header_row = 5
    headers = ["Obs No.", "Date", "Area/Equipment", "Description", "Impact", "Likelihood", "Severity", "Risk Rating", "Corrective Action", "Deadline", "Photo Evidence", "Closed Photo", "Status"]
    for col_idx, header_text in enumerate(headers, 1):
        cell = sheet.cell(row=header_row, column=col_idx)
        cell.value = header_text; cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    col_widths = {'A': 8, 'B': 20, 'C': 35, 'D': 60, 'E': 45, 'F': 12, 'G': 12, 'H': 12, 'I': 60, 'J': 15, 'K': EXCEL_PHOTO_COLUMN_WIDTH_UNITS, 'L': 20, 'M': 12}
    for col_letter, width in col_widths.items(): sheet.column_dimensions[col_letter].width = width

    for entry in observations_data:
        risk_rating = entry.get('risk_rating', 0)
        row_values = [
            entry.get('display_id'), entry.get('date_str'), entry.get('area_equipment'),
            entry.get('description'), entry.get('impact'), entry.get('likelihood'),
            entry.get('severity'), risk_rating, entry.get('corrective_action'),
            entry.get('deadline'), None, "Attach closed photo", "Open"
        ]
        sheet.append(row_values)
        new_row_num = sheet.max_row
        sheet.row_dimensions[new_row_num].height = EXCEL_ROW_HEIGHT_FOR_PHOTO_PT
        for col_idx_loop in range(1, len(headers) + 1):
            sheet.cell(row=new_row_num, column=col_idx_loop).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        risk_cell = sheet.cell(row=new_row_num, column=headers.index("Risk Rating") + 1)
        if 1 <= risk_rating <= 4: risk_cell.fill = green_fill
        elif 5 <= risk_rating <= 9: risk_cell.fill = yellow_fill
        elif 10 <= risk_rating <= 15: risk_cell.fill = orange_fill
        elif risk_rating >= 16: risk_cell.fill = red_fill
        photo_bytes_data = entry.get('photo_bytes')
        if photo_bytes_data:
            try:
                img = OpenpyxlImage(io.BytesIO(photo_bytes_data))
                img.width, img.height = EXCEL_PHOTO_TARGET_WIDTH_PX, EXCEL_PHOTO_TARGET_HEIGHT_PX
                sheet.add_image(img, get_column_letter(headers.index('Photo Evidence') + 1) + str(new_row_num))
            except Exception as e:
                print(f"Error embedding photo: {e}")
                sheet.cell(row=new_row_num, column=(headers.index('Photo Evidence') + 1)).value = "Error"
        else:
            sheet.cell(row=new_row_num, column=(headers.index('Photo Evidence') + 1)).value = "No Photo"

    excel_stream = io.BytesIO(); workbook.save(excel_stream); excel_stream.seek(0)
    return excel_stream

def register_callbacks(app):
    """Registers all callbacks for the observation app."""
    @app.callback(
        Output('manual-input-fields', 'style'),
        Output('observation-label', 'children'),
        Input('ai-toggle-switch', 'value')
    )
    def toggle_manual_fields(ai_switch_value):
        if ai_switch_value:
            return {'display': 'none'}, "Observation Details:"
        else:
            return {'display': 'block'}, "Description:"

    @app.callback(
        Output('manual-risk-rating-input', 'value'),
        Input('manual-likelihood-input', 'value'),
        Input('manual-severity-input', 'value')
    )
    def calculate_risk_rating(likelihood, severity):
        if likelihood and severity:
            return likelihood * severity
        return None

    @app.callback(
        Output('industry-update-status', 'children'),
        Input('industry-dropdown', 'value'),
        prevent_initial_call=True
    )
    def update_user_industry(selected_industry):
        if not current_user.is_authenticated or not selected_industry:
            raise PreventUpdate
        if selected_industry != current_user.industry:
            try:
                user_management.update_user_industry(current_user.id, selected_industry)
                current_user.industry = selected_industry
                return dbc.Alert(f"Your default industry has been updated to {selected_industry}.",
                                 color="success", duration=4000, fade=True,
                                 style={'position': 'fixed', 'top': '80px', 'right': '20px', 'zIndex': '9999'})
            except Exception as e:
                return dbc.Alert(f"Error updating industry: {e}", color="danger", duration=4000, fade=True)
        return no_update

    @app.callback(
        Output('flash-messages-container', 'children'),
        # FIX: Add outputs to clear all form fields upon successful submission
        Output('area-equipment-input', 'value'),
        Output('observation-textarea', 'value'),
        Output('photo-upload', 'contents'),
        Output('selected-file-name', 'children', allow_duplicate=True),
        Output('manual-impact-input', 'value'),
        Output('manual-likelihood-input', 'value'),
        Output('manual-severity-input', 'value'),
        Output('manual-corrective-action-input', 'value'),
        Output('manual-deadline-input', 'value'),
        Input('add-button', 'n_clicks'),
        [
            State('ai-toggle-switch', 'value'),
            State('industry-dropdown', 'value'),
            State('area-equipment-input', 'value'),
            State('observation-textarea', 'value'),
            State('photo-upload', 'contents'),
            State('manual-impact-input', 'value'),
            State('manual-likelihood-input', 'value'),
            State('manual-severity-input', 'value'),
            State('manual-corrective-action-input', 'value'),
            State('manual-deadline-input', 'value'),
        ],
        prevent_initial_call=True
    )
    def add_observation(n_clicks, ai_switch_value, industry, area_equipment, observation, photo_contents,
                        manual_impact, manual_likelihood, manual_severity, manual_corrective_action, manual_deadline):
        if not current_user.is_authenticated: raise PreventUpdate

        if not all([industry, area_equipment, observation]):
            msg = dbc.Alert("Industry, Area/Equipment, and Description fields are required.", color="warning", duration=5000)
            return (msg,) + (no_update,) * 9

        is_ai_enabled = bool(ai_switch_value)
        observation_data = {}

        if is_ai_enabled:
            ai_analysis = get_ai_analysis(observation, area_equipment, industry)
            observation_data = {
                'description': ai_analysis.get('CorrectedDescription'),
                'impact': ai_analysis.get('ImpactOnOperations'),
                'likelihood': ai_analysis.get('Likelihood'),
                'severity': ai_analysis.get('Severity'),
                'corrective_action': ai_analysis.get('CorrectiveAction'),
                'deadline': ai_analysis.get('DeadlineSuggestion'),
            }
        else:
            if not all([manual_impact, manual_likelihood, manual_severity, manual_corrective_action, manual_deadline]):
                msg = dbc.Alert("When AI is off, all manual fields are required.", color="warning", duration=5000)
                return (msg,) + (no_update,) * 9

            observation_data = {
                'description': observation,
                'impact': manual_impact,
                'likelihood': manual_likelihood,
                'severity': manual_severity,
                'corrective_action': manual_corrective_action,
                'deadline': manual_deadline,
            }

        observation_data['date_str'] = datetime.datetime.now().strftime("%d-%b-%Y")
        observation_data['area_equipment'] = area_equipment
        observation_data['photo_bytes'] = None
        if photo_contents:
            content_type, content_string = photo_contents.split(',')
            decoded_bytes = base64.b64decode(content_string)
            observation_data['photo_bytes'] = compress_image(decoded_bytes)

        db_observations.add_observation_to_db(observation_data, current_user.id)
        
        success_message = dbc.Alert("Observation successfully added to the database.", color="success", duration=5000)
        # FIX: Return empty values to clear the form fields
        return success_message, '', '', None, '', '', None, None, '', ''

    @app.callback(
        Output('report-content-container', 'children'),
        Input('url', 'pathname'),
        Input('search-input', 'value'),
        Input('sort-dropdown', 'value'),
        Input('store-refresh-signal', 'data')
    )
    def update_report_view(pathname, search_term, sort_by, refresh_signal):
        if pathname != '/report' or not current_user.is_authenticated: raise PreventUpdate
        observations = db_observations.get_observations_from_db(current_user.id, search_term, sort_by)
        if not observations: return html.P("No observations found. Add your first one!", style={'textAlign': 'center', 'padding': '50px'})
        cards = [design_components.create_observation_card(obs) for obs in observations]
        return cards

    @app.callback(
        Output('download-excel', 'data'),
        Input('download-report-button', 'n_clicks'),
        prevent_initial_call=True
    )
    def download_full_report(n_clicks):
        if not current_user.is_authenticated: raise PreventUpdate
        obs_for_excel = db_observations.get_observations_from_db(current_user.id, sort_by='date_oldest')
        if not obs_for_excel: raise PreventUpdate
        excel_stream = generate_excel_for_download(obs_for_excel)
        safe_email_part = current_user.email.split('@')[0]
        filename = f"Safety_Report_{safe_email_part}_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
        return dcc.send_bytes(excel_stream.read(), filename)

    @app.callback(
        Output('selected-file-name', 'children', allow_duplicate=True),
        Input('photo-upload', 'filename'),
        prevent_initial_call=True
    )
    def update_filename_display(filename):
        return f"File selected: {filename}" if filename else ""

    @app.callback(
        Output('confirm-delete-dialog', 'displayed'),
        Output('store-id-to-delete', 'data'),
        Input({'type': 'delete-button', 'index': ALL}, 'n_clicks'),
        prevent_initial_call=True
    )
    def display_delete_confirmation(n_clicks_list):
        if not any(n > 0 for n in n_clicks_list if n is not None): raise PreventUpdate
        ctx = dash.callback_context
        if not ctx.triggered_id: raise PreventUpdate
        obs_id_to_delete = ctx.triggered_id['index']
        return True, obs_id_to_delete

    @app.callback(
        Output('store-refresh-signal', 'data'),
        Output('delete-status-message', 'children'),
        Input('confirm-delete-dialog', 'submit_n_clicks'),
        State('store-id-to-delete', 'data'),
        State('store-refresh-signal', 'data'),
        prevent_initial_call=True
    )
    def process_deletion(submit_n_clicks, obs_id, refresh_count):
        if not obs_id or not current_user.is_authenticated: return no_update, no_update
        try:
            db_observations.delete_observation_from_db(obs_id, current_user.id)
            message = html.Div("Observation successfully deleted.", className="message-success")
            return refresh_count + 1, message
        except PermissionError as e:
             message = html.Div(f"Error: {e}", className="message-error")
             return no_update, message
        except Exception as e:
            print(f"Error during deletion of observation {obs_id}: {e}")
            message = html.Div(f"Error: Could not delete observation #{obs_id}.", className="message-error")
            return no_update, message