# gate_pass_app.py

import dash
from dash import dcc, html, Input, Output, State, no_update, ALL
import dash_bootstrap_components as dbc
from dash.exceptions import PreventUpdate
from flask_login import current_user
import datetime
import io
import os
import base64
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from psycopg2.errors import UniqueViolation

import db_gate_pass
import design_components
from utils import compress_image

def create_gate_pass_form_layout():
    """Creates the Gate Pass entry form."""
    return html.Div(className="observation-form-container", style={'maxWidth': '800px', 'margin': 'auto'}, children=[
        html.Div(id='gp-form-message'),
        html.Div(className="form-content", children=[
            html.Div([
                html.H1("Gate Pass Registration", className="form-title", style={'marginBottom': '0'}),
                dcc.Link(dbc.Button("View Report", color="info"), href="/gate-pass/report")
            ], style={'display': 'flex', 'justifyContent': 'space-between', 'alignItems': 'center', 'marginBottom': '35px'}),

            dbc.Row([
                dbc.Col(html.Div(className="form-group", children=[
                    dbc.Label("Date Issued"),
                    dcc.DatePickerSingle(id='gp-date-issued', date=datetime.date.today(), style={'width': '100%'})
                ]), md=6),
                dbc.Col(html.Div(className="form-group", children=[
                    dbc.Label("Gate Pass Number"),
                    dbc.Input(id='gp-gate-pass-number', placeholder="Enter unique pass number")
                ]), md=6),
            ]),
            html.Div(className="form-group", children=[
                dbc.Label("Item Description"),
                dbc.Textarea(id='gp-item-description', placeholder="Detailed description of the item(s)...", rows=4)
            ]),
            dbc.Row([
                dbc.Col(html.Div(className="form-group", children=[
                    dbc.Label("Issued To"),
                    dbc.Input(id='gp-issued-to', placeholder="Name of person receiving item")
                ]), md=6),
                dbc.Col(html.Div(className="form-group", children=[
                    dbc.Label("Company"),
                    dbc.Input(id='gp-company', placeholder="e.g., Contractor Company Name")
                ]), md=6),
            ]),
             dbc.Row([
                dbc.Col(html.Div(className="form-group", children=[
                    dbc.Label("Purpose of Removal"),
                    dbc.Input(id='gp-purpose')
                ]), md=6),
                dbc.Col(html.Div(className="form-group", children=[
                    dbc.Label("Type"),
                    dcc.Dropdown(id='gp-type', options=['returnable', 'non returnable'], value='returnable', clearable=False)
                ]), md=6),
            ]),
             dbc.Row([
                dbc.Col(html.Div(className="form-group", children=[
                    dbc.Label("Authorized By"),
                    dbc.Input(id='gp-authorized-by', placeholder="e.g., Manager's Name")
                ]), md=6),
                dbc.Col(html.Div(className="form-group", children=[
                    dbc.Label("Authorizing Department"),
                    dbc.Input(id='gp-authorizing-department', placeholder="e.g., IT, Engineering")
                ]), md=6),
            ]),
            html.Div(className="form-group", children=[
                dbc.Label("Date to be Returned"),
                dcc.DatePickerSingle(id='gp-date-to-be-returned', style={'width': '100%'})
            ]),
            html.Div(className="form-group", children=[
                html.Label("Item Picture (Taken Out)"),
                dcc.Upload(id='gp-photo-taken-out', className="photo-upload-area-styled", children=html.Div([
                    html.Img(src='/assets/upload-icon.png', className="upload-icon-img-styled"),
                    html.Span("Upload Photo")
                ], className="photo-upload-button-styled"), multiple=False),
                html.Span(id="gp-selected-file-name")
            ]),
            html.Div(className="submit-button-container", children=[
                html.Button("Submit Gate Pass", id="gp-submit-button", n_clicks=0, className="submit-button-style")
            ])
        ])
    ])

def create_gate_pass_report_layout():
    """Creates the Gate Pass report page."""
    return html.Div([
        dcc.Download(id='download-gp-excel'),
        dcc.ConfirmDialog(id='gp-confirm-delete', message='Are you sure you want to delete this gate pass? This action is permanent.'),
        dcc.Store(id='gp-item-to-delete-id'),
        dcc.Store(id='gp-refresh-signal', data=0),
        
        html.Div(className="report-main-content", children=[
            html.H1("Gate Pass Report", className="form-title"),
            html.Div(id='gp-report-action-message'),
            html.Div(className="report-controls", children=[
                dcc.Input(id='gp-search-input', type='text', placeholder='Search by pass #, item, person...', debounce=True, className='search-bar'),
                dbc.Button("Download Report (XLSX)", id="gp-download-button", color="success", className="nav-download-button"),
                dcc.Link(dbc.Button("New Entry", color="primary"), href="/gate-pass", className="ms-2")
            ]),
            dcc.Loading(id="loading-gp-report", type="default", children=html.Div(id='gp-report-table-container'))
        ]),

        dbc.Modal([
            dbc.ModalHeader("Update Gate Pass Return Status"),
            dbc.ModalBody([
                dbc.Label("Status"),
                dcc.Dropdown(
                    id='edit-gp-status',
                    options=[
                        {'label': 'Returned', 'value': 'returned'},
                        {'label': 'Non Returned', 'value': 'non returned'},
                    ],
                    className="mb-3",
                    clearable=False
                ),
                dbc.Label("Returned Date"),
                dcc.DatePickerSingle(id='edit-gp-returned-date', className="mb-2", style={'width': '100%'}),
                dbc.Label("Received By"), dbc.Input(id='edit-gp-received-by', className="mb-2"),
                dbc.Label("Remarks"), dbc.Textarea(id='edit-gp-remarks', className="mb-3"),
                dbc.Label("Item Picture (Returned Back)"),
                dcc.Upload(
                    id='edit-gp-photo-returned-back',
                    children=html.Div(['Drag and Drop or ', html.A('Select an Image')]),
                    style={'width': '100%', 'height': '60px', 'lineHeight': '60px', 'borderWidth': '1px', 'borderStyle': 'dashed', 'borderRadius': '5px', 'textAlign': 'center'},
                ),
                html.Div(id='edit-gp-photo-filename')
            ]),
            dbc.ModalFooter([
                dbc.Button("Cancel", id="gp-edit-cancel-button", color="secondary"),
                dbc.Button("Save Changes", id="gp-edit-save-button", color="primary"),
            ])
        ], id='gp-edit-modal', is_open=False, scrollable=True),
        dcc.Store(id='gp-item-to-edit-id'),
        dcc.Store(id='gp-photo-returned-bytes-store'),
    ])

def generate_gate_pass_excel(items_data):
    """Generates an Excel report for Gate Pass records matching the specified format."""
    EXCEL_PHOTO_TARGET_WIDTH_PX = 150
    EXCEL_PHOTO_TARGET_HEIGHT_PX = 112
    EXCEL_ROW_HEIGHT_FOR_PHOTO_PT = 90.0
    EXCEL_PHOTO_COLUMN_WIDTH_UNITS = 22

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Gate Pass Records"

    # --- Header and Logo Section ---
    sheet.merge_cells('A1:C4')
    logo_path = os.path.join('assets', 'riskwatch-logo.png')
    if os.path.exists(logo_path):
        logo = OpenpyxlImage(logo_path)
        logo.height = 75; logo.width = 175
        sheet.add_image(logo, 'A1')

    sheet.merge_cells('D1:Q4') # Updated merge range
    title_cell = sheet['D1']
    title_cell.value = "gatepass record"
    title_cell.font = Font(name='Calibri', size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Column Headers ---
    header_row = 5
    headers = [
        "s. no.", "date issued", "gate pass number", "items description", "issued to", "company",
        "purpose of removal", "authorized by", "authorizing department", "type", "date to be returned",
        "item picture taken-out", "item picture returned back", "returned date", "received by", "status", "remarks"
    ]
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    for col_idx, header_text in enumerate(headers, 1):
        cell = sheet.cell(row=header_row, column=col_idx)
        cell.value = header_text.title()
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # --- Data Rows ---
    current_row = header_row
    for idx, item in enumerate(items_data, 1):
        current_row += 1
        date_issued_str = item['date_issued'].strftime('%d-%b-%y') if item.get('date_issued') else ''
        date_to_be_returned_str = item['date_to_be_returned'].strftime('%d-%b-%y') if item.get('date_to_be_returned') else ''
        returned_date_str = item['returned_date'].strftime('%d-%b-%y') if item.get('returned_date') else ''
        
        row_values = [
            idx, date_issued_str, item.get('gate_pass_number', ''), item.get('item_description', ''),
            item.get('issued_to', ''), item.get('company', ''), item.get('purpose_of_removal', ''),
            item.get('authorized_by', ''), item.get('authorizing_department', ''), item.get('type', ''),
            date_to_be_returned_str, None, None, returned_date_str, item.get('received_by', ''),
            item.get('status', ''), item.get('remarks', '')
        ]
        
        # Write data to cells
        for col_idx_loop, value in enumerate(row_values, 1):
            sheet.cell(row=current_row, column=col_idx_loop, value=value)

        sheet.row_dimensions[current_row].height = EXCEL_ROW_HEIGHT_FOR_PHOTO_PT
        for col in range(1, len(headers) + 1):
            sheet.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')

        # Embed 'taken-out' photo
        photo_out_bytes = item.get('item_picture_taken_out_bytes')
        if photo_out_bytes:
            try:
                img_out = OpenpyxlImage(io.BytesIO(photo_out_bytes))
                img_out.width, img_out.height = EXCEL_PHOTO_TARGET_WIDTH_PX, EXCEL_PHOTO_TARGET_HEIGHT_PX
                photo_col_letter = get_column_letter(headers.index('item picture taken-out') + 1)
                sheet.add_image(img_out, f"{photo_col_letter}{current_row}")
            except Exception as e:
                print(f"Error embedding 'taken-out' photo for item {item['id']}: {e}")

        # Embed 'returned-back' photo
        photo_ret_bytes = item.get('item_picture_returned_back_bytes')
        if photo_ret_bytes:
            try:
                img_ret = OpenpyxlImage(io.BytesIO(photo_ret_bytes))
                img_ret.width, img_ret.height = EXCEL_PHOTO_TARGET_WIDTH_PX, EXCEL_PHOTO_TARGET_HEIGHT_PX
                photo_col_letter = get_column_letter(headers.index('item picture returned back') + 1)
                sheet.add_image(img_ret, f"{photo_col_letter}{current_row}")
            except Exception as e:
                print(f"Error embedding 'returned-back' photo for item {item['id']}: {e}")

    # --- Column Widths ---
    col_widths = {'A': 8, 'B': 15, 'C': 18, 'D': 45, 'E': 25, 'F': 25, 'G': 25, 'H': 25, 'I': 25, 'J': 15, 'K': 18, 'L': EXCEL_PHOTO_COLUMN_WIDTH_UNITS, 'M': EXCEL_PHOTO_COLUMN_WIDTH_UNITS, 'N': 18, 'O': 25, 'P': 15, 'Q': 35}
    for col_letter, width in col_widths.items():
        sheet.column_dimensions[col_letter].width = width

    excel_stream = io.BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream


def register_callbacks(app):
    @app.callback(
        Output('gp-form-message', 'children'),
        # Add outputs to clear form on success
        *[Output(id, 'value') for id in ['gp-gate-pass-number', 'gp-item-description', 'gp-issued-to', 'gp-company', 'gp-purpose', 'gp-authorized-by', 'gp-authorizing-department']],
        Output('gp-date-issued', 'date'),
        Output('gp-date-to-be-returned', 'date'),
        Output('gp-photo-taken-out', 'contents'),
        Output('gp-selected-file-name', 'children', allow_duplicate=True),
        Input('gp-submit-button', 'n_clicks'),
        [
            # --- START OF FIX: Explicitly define States with correct properties ---
            State('gp-date-issued', 'date'),
            State('gp-gate-pass-number', 'value'),
            State('gp-item-description', 'value'),
            State('gp-issued-to', 'value'),
            State('gp-company', 'value'),
            State('gp-purpose', 'value'),
            State('gp-authorized-by', 'value'),
            State('gp-authorizing-department', 'value'),
            State('gp-type', 'value'),
            State('gp-date-to-be-returned', 'date'),
            State('gp-photo-taken-out', 'contents')
            # --- END OF FIX ---
        ],
        prevent_initial_call=True
    )
    def submit_gp_entry(n_clicks, date_issued, pass_num, desc, issued_to, company, purpose, auth_by, auth_dept, type, date_return, photo_contents):
        clear_outputs = [no_update] * 11
        if not all([date_issued, pass_num, desc, issued_to, company, purpose, auth_by, auth_dept, type]):
            return (dbc.Alert("Please fill all required fields.", color="warning", duration=4000), *clear_outputs)

        data = {
            'date_issued': date_issued, 'gate_pass_number': pass_num, 'item_description': desc,
            'issued_to': issued_to, 'company': company, 'purpose_of_removal': purpose,
            'authorized_by': auth_by, 'authorizing_department': auth_dept, 'type': type,
            'date_to_be_returned': date_return, 'item_picture_taken_out_bytes': None
        }

        if photo_contents:
            _, content_string = photo_contents.split(',')
            data['item_picture_taken_out_bytes'] = compress_image(base64.b64decode(content_string))

        try:
            db_gate_pass.add_gate_pass_record_to_db(data, current_user.id, current_user.company_id)
            return (
                dbc.Alert("Gate Pass added successfully!", color="success", duration=4000),
                '', '', '', '', '', '', '',  # Clear text inputs
                datetime.date.today(), None, None, ''  # Clear date, upload, and filename
            )
        except UniqueViolation:
            return (dbc.Alert(f"Error: The Gate Pass Number '{pass_num}' is already in use for your company.", color="danger"), *clear_outputs)
        except Exception as e:
            return (dbc.Alert(f"An unexpected error occurred: {e}", color="danger"), *clear_outputs)

    @app.callback(
        Output('gp-report-table-container', 'children'),
        Input('gp-refresh-signal', 'data'),
        Input('gp-search-input', 'value')
    )
    def update_gp_report(refresh_signal, search_term):
        if not current_user.is_authenticated: raise PreventUpdate
        items = db_gate_pass.get_gate_pass_records_from_db(current_user.id, search_term)
        if not items:
            return html.P("No gate pass records found.", className="text-center mt-4")
        return [design_components.create_gate_pass_card(item) for item in items]

    @app.callback(
        Output('gp-edit-modal', 'is_open'),
        Output('gp-item-to-edit-id', 'data'),
        Output('edit-gp-status', 'value'),
        Output('edit-gp-returned-date', 'date'),
        Output('edit-gp-received-by', 'value'),
        Output('edit-gp-remarks', 'value'),
        Input({'type': 'gp-edit-btn', 'index': ALL}, 'n_clicks'),
        State('gp-edit-modal', 'is_open'),
        prevent_initial_call=True
    )
    def toggle_edit_modal(n_clicks, is_open):
        ctx = dash.callback_context
        if not ctx.triggered_id or not any(n for n in n_clicks if n is not None): raise PreventUpdate
        item_id = ctx.triggered_id['index']
        item = db_gate_pass.get_gate_pass_record_by_id(item_id, current_user.id)
        if not item: return False, *(no_update for _ in range(5))
        return (
            not is_open, item_id, item.get('status', 'non returned'), item.get('returned_date'),
            item.get('received_by', ''), item.get('remarks', '')
        )

    @app.callback(
        Output('gp-photo-returned-bytes-store', 'data'),
        Output('edit-gp-photo-filename', 'children'),
        Input('edit-gp-photo-returned-back', 'contents'),
        State('edit-gp-photo-returned-back', 'filename'),
        prevent_initial_call=True
    )
    def store_edited_photo(contents, filename):
        if contents:
            compressed_bytes = compress_image(base64.b64decode(contents.split(',')[1]))
            b64_string = base64.b64encode(compressed_bytes).decode('utf-8')
            return b64_string, f"File selected: {filename}"
        return no_update, no_update

    @app.callback(
        Output('gp-report-action-message', 'children'),
        Output('gp-refresh-signal', 'data', allow_duplicate=True),
        Output('gp-edit-modal', 'is_open', allow_duplicate=True),
        Input('gp-edit-save-button', 'n_clicks'),
        [
            State('gp-item-to-edit-id', 'data'), State('gp-refresh-signal', 'data'),
            State('edit-gp-status', 'value'), State('edit-gp-returned-date', 'date'),
            State('edit-gp-received-by', 'value'), State('edit-gp-remarks', 'value'),
            State('gp-photo-returned-bytes-store', 'data')
        ],
        prevent_initial_call=True
    )
    def save_gp_edit(n_clicks, item_id, r_count, status, r_date, r_by, remarks, photo_b64):
        if not item_id: raise PreventUpdate
        photo_bytes = base64.b64decode(photo_b64) if photo_b64 else None
        data = {
            'status': status, 'returned_date': r_date, 'received_by': r_by,
            'remarks': remarks, 'item_picture_returned_back_bytes': photo_bytes
        }
        try:
            db_gate_pass.update_gate_pass_record_in_db(item_id, data, current_user.id)
            return dbc.Alert("Record updated!", "success", duration=3000), r_count + 1, False
        except Exception as e:
            return dbc.Alert(f"Error: {e}", "danger"), no_update, True

    @app.callback(
        Output('gp-edit-modal', 'is_open', allow_duplicate=True),
        Input('gp-edit-cancel-button', 'n_clicks'),
        prevent_initial_call=True
    )
    def cancel_gp_edit(n_clicks):
        return False

    @app.callback(
        Output('gp-confirm-delete', 'displayed'),
        Output('gp-item-to-delete-id', 'data'),
        Input({'type': 'gp-delete-btn', 'index': ALL}, 'n_clicks'),
        prevent_initial_call=True
    )
    def display_gp_delete_confirm(n_clicks):
        ctx = dash.callback_context
        if not ctx.triggered_id or not any(n for n in n_clicks if n is not None): raise PreventUpdate
        return True, ctx.triggered_id['index']

    @app.callback(
        Output('gp-report-action-message', 'children', allow_duplicate=True),
        Output('gp-refresh-signal', 'data', allow_duplicate=True),
        Input('gp-confirm-delete', 'submit_n_clicks'),
        State('gp-item-to-delete-id', 'data'),
        State('gp-refresh-signal', 'data'),
        prevent_initial_call=True
    )
    def process_gp_delete(submit_n_clicks, item_id, r_count):
        if not item_id: raise PreventUpdate
        try:
            db_gate_pass.delete_gate_pass_record_from_db(item_id, current_user.id)
            return dbc.Alert("Gate Pass deleted.", "success", duration=3000), r_count + 1
        except Exception as e:
            return dbc.Alert(f"Error: {e}", "danger"), no_update

    @app.callback(
        Output('download-gp-excel', 'data'),
        Input('gp-download-button', 'n_clicks'),
        prevent_initial_call=True
    )
    def download_gp_report(n_clicks):
        if not current_user.is_authenticated: raise PreventUpdate
        items = db_gate_pass.get_all_gp_records_for_user_role(current_user.id)
        if not items: raise PreventUpdate
        excel_stream = generate_gate_pass_excel(items)
        filename = f"GatePass_Report_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
        return dcc.send_bytes(excel_stream.read(), filename)
    
    @app.callback(
        Output('gp-selected-file-name', 'children', allow_duplicate=True),
        Input('gp-photo-taken-out', 'filename'),
        prevent_initial_call=True
    )
    def update_gp_filename_display(filename):
        return f"File selected: {filename}" if filename else ""