# lost_and_found_app.py

import dash
from dash import dcc, html, Input, Output, State, no_update, ALL
import dash_bootstrap_components as dbc
from dash.exceptions import PreventUpdate
from flask_login import current_user
import datetime
import io
import os
import base64
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from psycopg2.errors import UniqueViolation

import db_lost_and_found
import design_components
from utils import compress_image

def create_lost_and_found_form_layout():
    """Creates the Lost & Found entry form, mirroring the observation form's style."""
    return html.Div(className="observation-form-container", style={'maxWidth': '800px', 'margin': 'auto'}, children=[
        html.Div(id='lf-form-message'),
        dcc.Store(id='lf-refresh-signal', data=0),
        html.Div(className="form-content", children=[
            html.Div([
                html.H1("Lost & Found Registration", className="form-title", style={'marginBottom': '0'}),
                dcc.Link(dbc.Button("View Report", color="info"), href="/lost-and-found/report")
            ], style={'display': 'flex', 'justifyContent': 'space-between', 'alignItems': 'center', 'marginBottom': '35px'}),
            
            dbc.Row([
                dbc.Col(html.Div(className="form-group", children=[
                    dbc.Label("Date Found"),
                    dcc.DatePickerSingle(id='lf-date-found', date=datetime.date.today(), style={'width': '100%'})
                ]), md=6),
                dbc.Col(html.Div(className="form-group", children=[
                    dbc.Label("Time Found (HH:MM)"),
                    dbc.Input(id='lf-time-found', type='text', placeholder='e.g., 14:30')
                ]), md=6),
            ]),
            dbc.Row([
                dbc.Col(html.Div(className="form-group", children=[
                    dbc.Label("Ticket #"),
                    dbc.Input(id='lf-ticket-no', placeholder="Enter unique ticket number")
                ]), md=6),
                dbc.Col(html.Div(className="form-group", children=[
                    dbc.Label("Type of Item"),
                    dcc.Dropdown(id='lf-item-type', options=['Valuable', 'Non-Valuable', 'Perishable'], placeholder="Select type...")
                ]), md=6),
            ]),
            html.Div(className="form-group", children=[
                dbc.Label("Item Description"),
                dbc.Textarea(id='lf-item-description', placeholder="Detailed description of the item...", rows=4)
            ]),
            dbc.Row([
                dbc.Col(html.Div(className="form-group", children=[
                    dbc.Label("Location Found"),
                    dbc.Input(id='lf-location-found', placeholder="e.g., Room 204, Lobby")
                ]), md=6),
                dbc.Col(html.Div(className="form-group", children=[
                    dbc.Label("Found By"),
                    dbc.Input(id='lf-found-by', placeholder="Name of person who found it")
                ]), md=6),
            ]),
            dbc.Row([
                dbc.Col(html.Div(className="form-group", children=[
                    dbc.Label("Department"),
                    dbc.Input(id='lf-department', placeholder="e.g., Housekeeping, F&B")
                ]), md=6),
                dbc.Col(html.Div(className="form-group", children=[
                    dbc.Label("Received by (Security)"),
                    dbc.Input(id='lf-received-by', placeholder="Name of security personnel")
                ]), md=6),
            ]),
            html.Div(className="form-group", children=[
                dbc.Label("Stored In"),
                dbc.Input(id='lf-stored-in', placeholder="e.g., Security L&F Store")
            ]),
            html.Div(className="form-group", children=[
                html.Label("Attach Photo (Optional):"),
                dcc.Upload(id='lf-photo-upload', className="photo-upload-area-styled", children=html.Div([
                    html.Img(src='/assets/upload-icon.png', className="upload-icon-img-styled"),
                    html.Span("Upload Photo")
                ], className="photo-upload-button-styled"), multiple=False),
                html.Span(id="lf-selected-file-name")
            ]),
            html.Div(className="submit-button-container", children=[
                html.Button("Submit Entry", id="lf-submit-button", n_clicks=0, className="submit-button-style")
            ])
        ])
    ])

def create_lost_and_found_report_layout():
    """Creates the Lost & Found report page, mirroring the observation report's style."""
    return html.Div([
        dcc.Download(id='download-lf-excel'),
        dcc.ConfirmDialog(id='lf-confirm-delete', message='Are you sure you want to delete this item? This action is permanent.'),
        dcc.Store(id='lf-item-to-delete-id'),
        dcc.Store(id='lf-refresh-signal', data=0),
        
        html.Div(className="report-main-content", children=[
            html.H1("Lost & Found Report", className="form-title"),
            html.Div(id='lf-report-action-message'),
            html.Div(className="report-controls", children=[
                dcc.Input(id='lf-search-input', type='text', placeholder='Search by ticket, description, location...', debounce=True, className='search-bar'),
                dbc.Button("Download Report (XLSX)", id="lf-download-button", color="success", className="nav-download-button"),
                dcc.Link(dbc.Button("New Entry", color="primary"), href="/lost-and-found", className="ms-2")
            ]),
            dcc.Loading(id="loading-lf-report", type="default", children=html.Div(id='lf-report-table-container'))
        ]),

        dbc.Modal([
            dbc.ModalHeader("Edit Lost & Found Record"),
            dbc.ModalBody([
                # --- FIX: Add new Status dropdown field ---
                dbc.Label("Status"),
                dcc.Dropdown(
                    id='edit-status',
                    options=[
                        {'label': 'Unclaimed', 'value': 'Unclaimed'},
                        {'label': 'Claimed', 'value': 'Claimed'},
                        {'label': 'Handed over to police', 'value': 'Handed over to police'},
                        {'label': 'Disposed', 'value': 'Disposed'}
                    ],
                    className="mb-3",
                    clearable=False
                ),
                # --- End of fix ---
                dbc.Label("Owner ID No."), dbc.Input(id='edit-owner-id', className="mb-2"),
                dbc.Label("Claimer / Receiver / Disposer"), dbc.Input(id='edit-claimer', className="mb-2"),
                dbc.Label("Receiver Contact No."), dbc.Input(id='edit-contact-no', className="mb-2"),
                dbc.Label("Claim Date"), dcc.DatePickerSingle(id='edit-claim-date', className="mb-2", style={'width': '100%'}),
                dbc.Label("Claim Time (HH:MM)"), dbc.Input(id='edit-claim-time', className="mb-2"),
                dbc.Label("Handed Over By"), dbc.Input(id='edit-handed-over-by', className="mb-2"),
                dbc.Label("Remarks"), dbc.Textarea(id='edit-remarks'),
            ]),
            dbc.ModalFooter([
                dbc.Button("Cancel", id="lf-edit-cancel-button", color="secondary"),
                dbc.Button("Save Changes", id="lf-edit-save-button", color="primary"),
            ])
        ], id='lf-edit-modal', is_open=False, scrollable=True),
        dcc.Store(id='lf-item-to-edit-id'),
    ])

# --- FIX: Complete rewrite of the Excel generation function ---
def generate_lost_and_found_excel(items_data):
    """Generates an Excel report matching the new, specified format."""
    EXCEL_PHOTO_TARGET_WIDTH_PX = 150
    EXCEL_PHOTO_TARGET_HEIGHT_PX = 112
    EXCEL_ROW_HEIGHT_FOR_PHOTO_PT = 90.0
    EXCEL_PHOTO_COLUMN_WIDTH_UNITS = 22
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Lost and Found Records"

    # --- Header and Logo Section ---
    sheet.merge_cells('A1:C4')
    logo_path = os.path.join('assets', 'riskwatch-logo.png')
    if os.path.exists(logo_path):
        logo = OpenpyxlImage(logo_path)
        logo.height = 75
        logo.width = 175
        sheet.add_image(logo, 'A1')

    sheet.merge_cells('D1:S4')
    title_cell = sheet['D1']
    title_cell.value = "Lost and found records"
    title_cell.font = Font(name='Calibri', size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Column Headers ---
    header_row = 5
    headers = [
        "Date", "Time", "Ticket #", "Type of Item", "Item Description", "Location Found",
        "Found By", "Department", "Received by the Person from Security", "Stored In", "Photo",
        "status", "Owner ID No.", "Claimer / Receiver / Disposer", "Receiver Contact no.",
        "Date", "Time", "Handed over by", "Remarks"
    ]
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    for col_idx, header_text in enumerate(headers, 1):
        cell = sheet.cell(row=header_row, column=col_idx)
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # --- Data Rows ---
    current_row = header_row
    for item in items_data:
        current_row += 1
        claim_date_str = item['claim_date'].strftime('%d-%b-%y') if item.get('claim_date') else ''
        entry_date_str = item['entry_date'].strftime('%d-%b-%y') if item.get('entry_date') else ''
        row_values = [
            entry_date_str, item.get('entry_time', ''), item.get('ticket_no', ''),
            item.get('item_type', ''), item.get('item_description', ''), item.get('location_found', ''),
            item.get('found_by', ''), item.get('department', ''), item.get('received_by_security', ''),
            item.get('stored_in', ''), None,  # Placeholder for Photo
            item.get('status', 'Unclaimed'), # status
            item.get('owner_id_no', ''), item.get('claimer_receiver_disposer', ''),
            item.get('receiver_contact_no', ''), claim_date_str, item.get('claim_time', ''),
            item.get('handed_over_by', ''), item.get('remarks', '')
        ]
        
        for col_idx_loop, value in enumerate(row_values, 1):
            sheet.cell(row=current_row, column=col_idx_loop, value=value)

        sheet.row_dimensions[current_row].height = EXCEL_ROW_HEIGHT_FOR_PHOTO_PT
        
        for col_idx_loop in range(1, len(headers) + 1):
            sheet.cell(row=current_row, column=col_idx_loop).alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
        
        photo_bytes_data = item.get('photo_bytes')
        if photo_bytes_data:
            try:
                img = OpenpyxlImage(io.BytesIO(photo_bytes_data))
                img.width, img.height = EXCEL_PHOTO_TARGET_WIDTH_PX, EXCEL_PHOTO_TARGET_HEIGHT_PX
                photo_col_letter = get_column_letter(headers.index('Photo') + 1)
                sheet.add_image(img, f"{photo_col_letter}{current_row}")
            except Exception as e:
                print(f"Error embedding photo for item {item['id']}: {e}")
                sheet.cell(row=current_row, column=(headers.index('Photo') + 1)).value = "Error"

    # --- Column Widths ---
    col_widths = {
        'A': 12, 'B': 10, 'C': 12, 'D': 15, 'E': 45, 'F': 25, 'G': 18, 'H': 18,
        'I': 25, 'J': 18, 'K': EXCEL_PHOTO_COLUMN_WIDTH_UNITS, 'L': 22, 'M': 18,
        'N': 28, 'O': 22, 'P': 12, 'Q': 10, 'R': 18, 'S': 35
    }
    for col_letter, width in col_widths.items():
        sheet.column_dimensions[col_letter].width = width

    excel_stream = io.BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream
# --- End of Excel function rewrite ---


def register_callbacks(app):
    @app.callback(
        Output('lf-form-message', 'children'),
        Output('lf-refresh-signal', 'data', allow_duplicate=True),
        Output('lf-date-found', 'date'),
        Output('lf-time-found', 'value'),
        Output('lf-ticket-no', 'value'),
        Output('lf-item-type', 'value'),
        Output('lf-item-description', 'value'),
        Output('lf-location-found', 'value'),
        Output('lf-found-by', 'value'),
        Output('lf-department', 'value'),
        Output('lf-received-by', 'value'),
        Output('lf-stored-in', 'value'),
        Output('lf-photo-upload', 'contents'),
        Output('lf-selected-file-name', 'children', allow_duplicate=True),
        Input('lf-submit-button', 'n_clicks'),
        [
            State('lf-refresh-signal', 'data'),
            State('lf-date-found', 'date'), State('lf-time-found', 'value'),
            State('lf-ticket-no', 'value'), State('lf-item-type', 'value'),
            State('lf-item-description', 'value'), State('lf-location-found', 'value'),
            State('lf-found-by', 'value'), State('lf-department', 'value'),
            State('lf-received-by', 'value'), State('lf-stored-in', 'value'),
            State('lf-photo-upload', 'contents')
        ],
        prevent_initial_call=True
    )
    def submit_lf_entry(n_clicks, refresh_count, date, time, ticket, item_type, desc, location, found_by, dept, received_by, stored_in, photo_contents):
        if not all([date, time, ticket, item_type, desc, location, found_by, dept, stored_in]):
            return (dbc.Alert("Please fill all fields.", color="warning", duration=4000), *(no_update for _ in range(13)))

        data = {
            'entry_date': date, 'entry_time': time, 'ticket_no': ticket,
            'item_type': item_type, 'item_description': desc, 'location_found': location,
            'found_by': found_by, 'department': dept, 'received_by_security': received_by,
            'stored_in': stored_in, 'photo_bytes': None
        }
        
        if photo_contents:
            content_type, content_string = photo_contents.split(',')
            decoded_bytes = base64.b64decode(content_string)
            data['photo_bytes'] = compress_image(decoded_bytes)

        try:
            db_lost_and_found.add_lost_and_found_item_to_db(data, current_user.id, current_user.company_id)
            return (
                dbc.Alert("Entry added successfully!", color="success", duration=4000),
                refresh_count + 1,
                datetime.date.today(), '', '', None, '', '', '', '', '', '', None, ''
            )
        except UniqueViolation:
            return (dbc.Alert(f"Error: The Ticket # '{ticket}' is already in use for your company. Please enter a unique ticket number.", color="danger"), *(no_update for _ in range(13)))
        except Exception as e:
            return (dbc.Alert(f"An unexpected error occurred: {e}", color="danger"), *(no_update for _ in range(13)))

    @app.callback(
        Output('lf-report-table-container', 'children'),
        Input('url', 'pathname'),
        Input('lf-refresh-signal', 'data'),
        Input('lf-search-input', 'value')
    )
    def update_lf_report(pathname, refresh_signal, search_term):
        if pathname != '/lost-and-found/report':
            raise PreventUpdate

        if not current_user.is_authenticated: raise PreventUpdate
        items = db_lost_and_found.get_lost_and_found_items_from_db(current_user.id, search_term)
        
        if not items:
            return html.P("No records found for your company.", className="text-center mt-4")

        cards = [design_components.create_lost_and_found_card(item) for item in items]
        return cards

    # --- FIX: Update edit modal callback to handle the 'status' field ---
    @app.callback(
        Output('lf-edit-modal', 'is_open'),
        Output('lf-item-to-edit-id', 'data'),
        Output('edit-status', 'value'),
        Output('edit-owner-id', 'value'),
        Output('edit-claimer', 'value'),
        Output('edit-contact-no', 'value'),
        Output('edit-claim-date', 'date'),
        Output('edit-claim-time', 'value'),
        Output('edit-handed-over-by', 'value'),
        Output('edit-remarks', 'value'),
        Input({'type': 'lf-edit-btn', 'index': ALL}, 'n_clicks'),
        State('lf-edit-modal', 'is_open'),
        prevent_initial_call=True
    )
    def toggle_edit_modal(n_clicks, is_open):
        ctx = dash.callback_context
        if not ctx.triggered_id or not any(n for n in n_clicks if n is not None):
            raise PreventUpdate
        
        item_id = ctx.triggered_id['index']
        item = db_lost_and_found.get_lost_and_found_item_by_id(item_id, current_user.id)
        if not item:
            return False, *(no_update for _ in range(9))
        
        return (
            not is_open, item_id,
            item.get('status', 'Unclaimed'), # Populate status dropdown
            item.get('owner_id_no', ''), item.get('claimer_receiver_disposer', ''),
            item.get('receiver_contact_no', ''), item.get('claim_date'),
            item.get('claim_time', ''), item.get('handed_over_by', ''),
            item.get('remarks', '')
        )
    
    # --- FIX: Update save callback to include 'status' ---
    @app.callback(
        Output('lf-report-action-message', 'children', allow_duplicate=True),
        Output('lf-refresh-signal', 'data', allow_duplicate=True),
        Output('lf-edit-modal', 'is_open', allow_duplicate=True),
        Input('lf-edit-save-button', 'n_clicks'),
        [
            State('lf-item-to-edit-id', 'data'), State('lf-refresh-signal', 'data'),
            State('edit-status', 'value'), # Get status from dropdown
            State('edit-owner-id', 'value'), State('edit-claimer', 'value'),
            State('edit-contact-no', 'value'), State('edit-claim-date', 'date'),
            State('edit-claim-time', 'value'), State('edit-handed-over-by', 'value'),
            State('edit-remarks', 'value')
        ],
        prevent_initial_call=True
    )
    def save_lf_edit(n_clicks, item_id, refresh_count, status, owner_id, claimer, contact, claim_date, claim_time, handed_by, remarks):
        if not item_id: raise PreventUpdate
        
        data = {
            'status': status, # Include status in data payload
            'owner_id_no': owner_id, 'claimer_receiver_disposer': claimer,
            'receiver_contact_no': contact, 'claim_date': claim_date,
            'claim_time': claim_time, 'handed_over_by': handed_by, 'remarks': remarks
        }
        try:
            db_lost_and_found.update_lost_and_found_item_in_db(item_id, data, current_user.id)
            return dbc.Alert("Record updated!", color="success", duration=3000), refresh_count + 1, False
        except Exception as e:
            return dbc.Alert(f"Error: {e}", color="danger"), no_update, True

    @app.callback(
        Output('lf-edit-modal', 'is_open', allow_duplicate=True),
        Input('lf-edit-cancel-button', 'n_clicks'),
        prevent_initial_call=True
    )
    def cancel_edit(n_clicks):
        return False

    @app.callback(
        Output('lf-confirm-delete', 'displayed'),
        Output('lf-item-to-delete-id', 'data'),
        Input({'type': 'lf-delete-btn', 'index': ALL}, 'n_clicks'),
        prevent_initial_call=True
    )
    def display_delete_confirm(n_clicks):
        ctx = dash.callback_context
        if not ctx.triggered_id or not any(n for n in n_clicks if n is not None):
            raise PreventUpdate
        item_id = ctx.triggered_id['index']
        return True, item_id

    @app.callback(
        Output('lf-report-action-message', 'children', allow_duplicate=True),
        Output('lf-refresh-signal', 'data', allow_duplicate=True),
        Input('lf-confirm-delete', 'submit_n_clicks'),
        State('lf-item-to-delete-id', 'data'),
        State('lf-refresh-signal', 'data'),
        prevent_initial_call=True
    )
    def process_lf_delete(submit_n_clicks, item_id, refresh_count):
        if not item_id: raise PreventUpdate
        try:
            db_lost_and_found.delete_lost_and_found_item_from_db(item_id, current_user.id)
            return dbc.Alert("Item deleted.", color="success", duration=3000), refresh_count + 1
        except Exception as e:
            return dbc.Alert(f"Error: {e}", color="danger"), no_update

    @app.callback(
        Output('download-lf-excel', 'data'),
        Input('lf-download-button', 'n_clicks'),
        prevent_initial_call=True
    )
    def download_lf_report(n_clicks):
        if not current_user.is_authenticated: raise PreventUpdate
        items = db_lost_and_found.get_all_lf_items_for_user_role(current_user.id)
        if not items:
            raise PreventUpdate
        
        excel_stream = generate_lost_and_found_excel(items)
        filename = f"LostAndFound_Report_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
        return dcc.send_bytes(excel_stream.read(), filename)
    
    @app.callback(
        Output('lf-selected-file-name', 'children', allow_duplicate=True),
        Input('lf-photo-upload', 'filename'),
        prevent_initial_call=True
    )
    def update_filename_display(filename):
        return f"File selected: {filename}" if filename else "" 